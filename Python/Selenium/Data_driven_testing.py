import time

from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from openpyxl import Workbook
from selenium.webdriver.common.by import By

# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC


# Load the Excel sheet:
workbook = load_workbook('saucedemo_test_data.xlsx')

# Selecting the active sheet:
sheet = workbook.active

browser = webdriver.Chrome()
browser.maximize_window()

time.sleep(4)

# Iterating the value by rows:

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    username = row[0]
    password = row[1]

    url = "https://www.saucedemo.com/"
    browser.get(url)
    time.sleep(4)
    browser.find_element(By.ID, "user-name").send_keys(username)
    time.sleep(1)
    browser.find_element(By.ID,"password").send_keys(password)
    time.sleep(1)
    browser.find_element(By.ID, "login-button").click()

    time.sleep(4)
    browser.find_element(By.XPATH, '//button[@id="react-burger-menu-btn"]').click()

    time.sleep(4)
    browser.find_element(By.XPATH, '//a[@id="logout_sidebar_link"]').click()
    time.sleep(4)

browser.quit()