import time
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from openpyxl import Workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Load the Excel sheet:
workbook = load_workbook('saucedemo_test_data.xlsx')

# Selecting the active sheet:
sheet = workbook.active

browser = webdriver.Chrome()
browser.maximize_window()

# Wait for up to 10 seconds for elements to load
wait = WebDriverWait(browser, 10)

# Iterating through the rows:
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    username = row[0]
    password = str(row[1])  # ✅ Convert password to a string

    url = "https://www.saucedemo.com/"
    browser.get(url)
    time.sleep(2)  # Short sleep to ensure page is fully loaded

    # Wait for the username field to be present and enter login details
    wait.until(EC.presence_of_element_located((By.ID, "user-name"))).send_keys(username)
    browser.find_element(By.ID, "password").send_keys(password)
    browser.find_element(By.ID, "login-button").click()

    # **Verify login success by checking if the "Products" title is displayed**
    try:
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "title")))
        print(f"✅ Login successful for user: {username}")
    except:
        print(f"❌ Login failed for user: {username}")

    # Wait for the burger menu button to be clickable
    burger_menu_button = wait.until(EC.element_to_be_clickable((By.ID, "react-burger-menu-btn")))
    burger_menu_button.click()

    # Wait for the logout button to be clickable
    logout_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//a[@id="logout_sidebar_link"]')))
    logout_button.click()

    time.sleep(2)  # Wait before starting the next iteration

browser.quit()
